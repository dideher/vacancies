import logging
from django.utils.crypto import get_random_string
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension
from openpyxl.cell.cell import Cell
from openpyxl.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from sortedcontainers import SortedSet
from django.contrib.auth.models import User
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib import messages
from django.db.models import QuerySet, Count
from django.http import HttpResponse
from excel_response import ExcelResponse
from .forms import UploadFileForm
from vacancies.utils.permissions import check_user_is_superuser
from main_app.models import Entry, Specialty, EntryVariantType
from schools.models import School, SchoolGroup
from users.models import Profile
from history.models import HistoryEntry
from django.db import transaction

# Get an instance of a logger
logger = logging.getLogger(__name__)


def str2bool(value: str) -> bool:
    return value.lower() in ['true', '1', 't', 'y', 'yes', 'ναι']


class AggregatedEntriesReport:

    def __init__(self):
        started = datetime.now().replace(microsecond=0)
        self.entries: QuerySet[Entry] = Entry.objects.all().order_by("school__school_group__ordering", "school__id")

        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) fetched all entries in just '%s'", finished-started)

    def createSpecialtiesTypes(self):

        general_education_spc_types = list()
        special_education_spc_types = list()
        misc_spc_types = list()
        tmimata_entaksis_spc_types = list()
        paralili_stiriksi_spc_types = list()

        entries: QuerySet[dict] = Entry.objects.order_by('specialty__ordering').values('specialty__code',
                                                                                       'specialty__lectic',
                                                                                       'variant').annotate(Count('variant', distinct=True))
        for entry in entries:
            entry_variant = str(EntryVariantType(entry.get('variant')).label)
            entry_specialization = entry.get('specialty__code') + " " + entry.get('specialty__lectic')

            if entry_variant == 'Γενικής Αγωγής - Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                general_education_spc_types.append(f'{entry_specialization} - Γενικής Αγωγής - Πανελλαδικώς Εξεταζόμενα Μαθήματα')
                general_education_spc_types.append(f'{entry_specialization} - Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα')
                general_education_spc_types.append(f'{entry_specialization} - Γενικής Αγωγής (Σύνολο)')
            elif entry_variant == 'Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα':
                general_education_spc_types.append(f'{entry_specialization} - Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα')

            elif 'Τμήμα Ένταξης' in entry_variant:
                tmimata_entaksis_spc_types.append(f'{entry_specialization} - {entry_variant}')
            elif 'Παράλληλης Στήριξης' in entry_variant:
                paralili_stiriksi_spc_types.append(f'{entry_specialization} - {entry_variant}')
            elif 'Ειδικής Αγωγής' in entry_variant:
                special_education_spc_types.append(f'{entry_specialization} - {entry_variant}')
            else:
                misc_spc_types.append(f'{entry_specialization} - {entry_variant}')

        self.generalEducationSpcTypes = general_education_spc_types
        self.specialEducationSpcTypes = special_education_spc_types
        self.miscSpcTypes = misc_spc_types
        self.tmimata_entaksis_spc_types = tmimata_entaksis_spc_types
        self.paralili_stiriksi_spc_types = paralili_stiriksi_spc_types

    def getSchools(self):
        general_education_schools = list()
        special_education_schools = list()
        tmimata_entaksis_schools = list()
        paralili_stiriksi_schools = list()
        misc_schools = list()

        general_education_entries = dict()
        special_education_entries = dict()
        tmimata_entaksis_entries = dict()
        paralili_stiriksi_entries = dict()
        misc_entries = dict()

        for entry in self.entries:

            entry_variant = str(EntryVariantType(entry.variant).label)
            school = entry.school

            if entry_variant in ['Γενικής Αγωγής - Πανελλαδικώς Εξεταζόμενα Μαθήματα',
                                 'Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα']:
                school_entries: list[Entry] = general_education_entries.get(school.name, None)

                try:
                    school_entries.append(entry)
                except AttributeError:
                    general_education_entries[school.name] = [entry, ]

                if school not in general_education_schools:
                    general_education_schools.append(school)

            elif 'Τμήμα Ένταξης' in entry_variant:
                school_entries: list[Entry] = tmimata_entaksis_entries.get(school.name, None)

                try:
                    school_entries.append(entry)
                except AttributeError:
                    tmimata_entaksis_entries[school.name] = [entry, ]

                if school not in tmimata_entaksis_schools:
                    tmimata_entaksis_schools.append(school)
            elif 'Παράλληλης Στήριξης' in entry_variant:
                school_entries: list[Entry] = paralili_stiriksi_entries.get(school.name, None)
                try:
                    school_entries.append(entry)
                except AttributeError:
                    paralili_stiriksi_entries[school.name] = [entry, ]

                if school not in paralili_stiriksi_schools:
                    paralili_stiriksi_schools.append(school)
            elif 'Ειδικής Αγωγής' in entry_variant:
                school_entries: list[Entry] = special_education_entries.get(school.name, None)
                try:
                    school_entries.append(entry)
                except AttributeError:
                    special_education_entries[school.name] = [entry, ]

                if school not in special_education_schools:
                    special_education_schools.append(school)
            else:
                school_entries: list[Entry] = misc_entries.get(school.name, None)

                try:
                    school_entries.append(entry)
                except AttributeError:
                    misc_entries[school.name] = [entry, ]

                if school not in misc_schools:
                    misc_schools.append(school)

        self.generalEducationSchools = general_education_schools
        self.specialEducationSchools = special_education_schools
        self.miscSchools = misc_schools
        self.tmimata_entaksis_schools = tmimata_entaksis_schools
        self.paralili_stiriksi_schools = paralili_stiriksi_schools

        self.general_education_entries = general_education_entries
        self.special_education_entries = special_education_entries
        self.tmimata_entaksis_entries = tmimata_entaksis_entries
        self.paralili_stiriksi_entries = paralili_stiriksi_entries
        self.misc_entries = misc_entries

    def createTables(self):
        self.createGEStable()
        self.createSEStable()
        self.createMStable()

    def createMStable(self):
        self.msTable = list()
        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header += self.miscSpcTypes[:]

        self.msTable.append(header)
        for school in self.miscSchools:
            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                # TODO: Need to address timezone here
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')
            sch_values = [0] * len(self.miscSpcTypes)

            school_name = school.name

            for vacancy_entry in self.misc_entries.get(school_name):
                entry_variant = str(EntryVariantType(vacancy_entry.variant).label)
                entry_type = vacancy_entry.type
                entry_hours = vacancy_entry.hours
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'

                spc_type = f'{entry_specialization} - {entry_variant}'
                index_scp_type = self.miscSpcTypes.index(spc_type)

                if entry_type == 'Κενό':
                    sch_values[index_scp_type] -= int(entry_hours)
                else:
                    sch_values[index_scp_type] += int(entry_hours)

            entry += sch_values

            self.msTable.append(entry)

    def createSEStable(self):
        self.sesTable = list()
        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header += self.specialEducationSpcTypes[:]

        self.sesTable.append(header)
        for school in self.specialEducationSchools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')
            sch_values = [0] * len(self.specialEducationSpcTypes)

            school_name = school.name

            for vacancy_entry in self.special_education_entries.get(school_name):
                entry_variant = str(EntryVariantType(vacancy_entry.variant).label)
                entry_type = vacancy_entry.type
                entry_hours = vacancy_entry.hours
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'

                spc_type = f'{entry_specialization} - {entry_variant}'
                index_scp_type = self.specialEducationSpcTypes.index(spc_type)

                if entry_type == 'Κενό':
                    sch_values[index_scp_type] -= int(entry_hours)
                else:
                    sch_values[index_scp_type] += int(entry_hours)

            entry += sch_values

            self.sesTable.append(entry)

    def construct_tmimata_entakseis_table(self):
        return_table = list()
        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header += self.tmimata_entaksis_spc_types[:]

        return_table.append(header)
        for school in self.tmimata_entaksis_schools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')
            sch_values = [0] * len(self.tmimata_entaksis_spc_types)

            school_name = school.name

            for vacancy_entry in self.tmimata_entaksis_entries.get(school_name):
                entry_variant = str(EntryVariantType(vacancy_entry.variant).label)
                entry_type = vacancy_entry.type
                entry_hours = vacancy_entry.hours
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'

                spc_type = f'{entry_specialization} - {entry_variant}'
                index_scp_type = self.tmimata_entaksis_spc_types.index(spc_type)

                if entry_type == 'Κενό':
                    sch_values[index_scp_type] -= int(entry_hours)
                else:
                    sch_values[index_scp_type] += int(entry_hours)

            entry += sch_values

            return_table.append(entry)

        return return_table

    def construct_paralili_stiriksi_table(self):
        return_table = list()
        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header += self.paralili_stiriksi_spc_types[:]

        return_table.append(header)
        for school in self.paralili_stiriksi_schools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')
            sch_values = [0] * len(self.paralili_stiriksi_spc_types)

            school_name = school.name

            for vacancy_entry in self.paralili_stiriksi_entries.get(school_name):
                entry_variant = str(EntryVariantType(vacancy_entry.variant).label)
                entry_type = vacancy_entry.type
                entry_hours = vacancy_entry.hours
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'

                spc_type = f'{entry_specialization} - {entry_variant}'
                index_scp_type = self.paralili_stiriksi_spc_types.index(spc_type)

                if entry_type == 'Κενό':
                    sch_values[index_scp_type] -= int(entry_hours)
                else:
                    sch_values[index_scp_type] += int(entry_hours)

            entry += sch_values

            return_table.append(entry)

        return return_table

    def construct_ges_description_table(self):

        ges_description_table = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header.append('Παρατηρήσεις')

        ges_description_table.append(header)

        for school in self.generalEducationSchools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')

            school_entries_descriptions = list()

            school_name = school.name

            for vacancy_entry in self.general_education_entries.get(school_name):

                entry_type = vacancy_entry.type
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'
                entry_description: str = vacancy_entry.description

                if entry_description is not None and len(entry_description.strip()) > 0:
                    school_entries_descriptions.append(f'{entry_specialization} - {entry_type} - '
                                                       f'{entry_description}')

            entry.append("\n".join(school_entries_descriptions))

            ges_description_table.append(entry)

        return ges_description_table

    def construct_ses_description_table(self):

        ses_description_table = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header.append('Παρατηρήσεις')

        ses_description_table.append(header)

        for school in self.specialEducationSchools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')

            school_entries_descriptions = list()

            school_name = school.name

            for vacancy_entry in self.special_education_entries.get(school_name):

                entry_type = vacancy_entry.type
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'
                entry_description: str = vacancy_entry.description

                if entry_description is not None and len(entry_description.strip()) > 0:
                    school_entries_descriptions.append(f'{entry_specialization} - {entry_type} - '
                                                       f'{entry_description}')

            entry.append("\n".join(school_entries_descriptions))

            ses_description_table.append(entry)

        return ses_description_table

    def construct_misc_description_table(self):

        misc_description_table = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header.append('Παρατηρήσεις')

        misc_description_table.append(header)

        for school in self.miscSchools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')

            school_entries_descriptions = list()

            school_name = school.name

            for vacancy_entry in self.misc_entries.get(school_name):

                entry_type = vacancy_entry.type
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'
                entry_description: str = vacancy_entry.description

                if entry_description is not None and len(entry_description.strip()) > 0:
                    school_entries_descriptions.append(f'{entry_specialization} - {entry_type} - {entry_description}')

            entry.append("\n".join(school_entries_descriptions))

            misc_description_table.append(entry)

        return misc_description_table

    def construct_tmimata_entaksis_description_table(self):

        description_table = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header.append('Παρατηρήσεις')

        description_table.append(header)

        for school in self.tmimata_entaksis_schools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')

            school_entries_descriptions = list()

            school_name = school.name

            for vacancy_entry in self.tmimata_entaksis_entries.get(school_name):

                entry_type = vacancy_entry.type
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'
                entry_description: str = vacancy_entry.description

                if entry_description is not None and len(entry_description.strip()) > 0:
                    school_entries_descriptions.append(f'{entry_specialization} - {entry_type} - {entry_description}')

            entry.append("\n".join(school_entries_descriptions))

            description_table.append(entry)

        return description_table

    def construct_paralili_stiriksi_description_table(self):

        description_table = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header.append('Παρατηρήσεις')

        description_table.append(header)

        for school in self.paralili_stiriksi_schools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')

            school_entries_descriptions = list()

            school_name = school.name

            for vacancy_entry in self.paralili_stiriksi_entries.get(school_name):

                entry_type = vacancy_entry.type
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'
                entry_description: str = vacancy_entry.description

                if entry_description is not None and len(entry_description.strip()) > 0:
                    school_entries_descriptions.append(f'{entry_specialization} - {entry_type} - {entry_description}')

            entry.append("\n".join(school_entries_descriptions))

            description_table.append(entry)

        return description_table

    def createGEStable(self):

        self.gesTable = list()

        header = list()
        header.append('Ημ/νια Επικαιροποίησης')
        header.append('Σχολείο')
        header += self.generalEducationSpcTypes[:]

        self.gesTable.append(header)
        for school in self.generalEducationSchools:

            entry = list()

            if school.managed_by.status_time is None:
                entry.append("")
            else:
                entry.append(school.managed_by.status_time.strftime('%d/%m/%Y, %H:%M:%S'))

            entry.append(f'{school.school_group.name}, {school.name}')
            sch_values = [0] * len(self.generalEducationSpcTypes)

            school_name = school.name

            for vacancy_entry in self.general_education_entries.get(school_name):
                entry_variant = str(EntryVariantType(vacancy_entry.variant).label)
                entry_type = vacancy_entry.type
                entry_hours = vacancy_entry.hours
                entry_specialization = f'{vacancy_entry.specialty.code} {vacancy_entry.specialty.lectic}'

                spc_type = f'{entry_specialization} - {entry_variant}'
                index_scp_type = self.generalEducationSpcTypes.index(spc_type)

                if entry_type == 'Κενό':
                    sch_values[index_scp_type] -= int(entry_hours)
                else:
                    sch_values[index_scp_type] += int(entry_hours)

                spc_type = f'{entry_specialization} - Γενικής Αγωγής (Σύνολο)'
                if spc_type in self.generalEducationSpcTypes:
                    index_scp_type = self.generalEducationSpcTypes.index(spc_type)
                    if entry_type == 'Κενό':
                        sch_values[index_scp_type] -= int(entry_hours)
                    else:
                        sch_values[index_scp_type] += int(entry_hours)

            entry += sch_values

            self.gesTable.append(entry)

    def get_workbook(self):

        def filter_headers(header_row):
            # need to map / manipulate header values
            for idx, column in enumerate(header_row):
                if column.endswith('- Γενικής Αγωγής (Σύνολο)'):
                    header_row[idx] = header_row[idx].replace('- Γενικής Αγωγής (Σύνολο)', '- (Συν.)')
                elif column.endswith('- Γενικής Αγωγής - Πανελλαδικώς Εξεταζόμενα Μαθήματα'):
                    header_row[idx] = header_row[idx].replace(
                        '- Γενικής Αγωγής - Πανελλαδικώς Εξεταζόμενα Μαθήματα', '- Π.ΕΞ.')
                elif column.endswith('- Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα'):
                    header_row[idx] = header_row[idx].replace(
                        '- Γενικής Αγωγής - μη Πανελλαδικώς Εξεταζόμενα Μαθήματα', '')

        def style_description_ws(ws):
            # style header
            for col in range(ws.min_column, ws.max_column + 1):
                ws[get_column_letter(col) + '1'].style = 'header_style_description'

            # update column dimensions
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                if col == 1:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=22)
                elif col == 2:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=35)
                else:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=150)

            ws.column_dimensions = dim_holder

            # update 1st row height
            ws.row_dimensions[1] = RowDimension(ws, height=25)

            # update other rows, based on content of 3 column
            for row in range(ws.min_row+1, ws.max_row + 1):

                cell: Cell = ws[get_column_letter(3) + str(row)]

                # how many new lines we have in the cell
                count_of_crlf = str(cell.value).count('\n')
                ws.row_dimensions[row] = RowDimension(ws, height=14 * (count_of_crlf + 1))

            # update cells other than header
            for row in range(ws.min_row+1, ws.max_row + 1):
                for col in range(ws.min_column, ws.max_column + 1):
                    cell: Cell = ws[get_column_letter(col) + str(row)]
                    if col == 1:
                        cell.alignment = Alignment(vertical="top")
                    elif col == 2:
                        cell.alignment = Alignment(vertical="top")
                    elif col == 3:
                        cell.alignment = Alignment(wrapText=True, vertical="top")

        def style_ws(ws):

            # style header
            for col in range(ws.min_column, ws.max_column + 1):
                ws[get_column_letter(col) + '1'].style = 'header_style'

            # update column dimensions
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                if col == 1:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=22)
                elif col == 2:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=35)
                else:
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=7)

            ws.column_dimensions = dim_holder
            # update 1st row height
            ws.row_dimensions[1] = RowDimension(ws, height=150)

        started = datetime.now().replace(microsecond=0)
        self.getSchools()
        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) completed getSchools() in just '%s'", finished - started)

        started = datetime.now().replace(microsecond=0)
        self.createSpecialtiesTypes()
        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) completed createSpecialtiesTypes() in just '%s'", finished - started)

        started = datetime.now().replace(microsecond=0)
        self.createTables()
        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) completed createTables() in just '%s'", finished - started)

        started = datetime.now().replace(microsecond=0)
        tmimata_entakseis_table = self.construct_tmimata_entakseis_table()
        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) completed construct_tmimata_entakseis_table in just '%s'", finished - started)

        started = datetime.now().replace(microsecond=0)
        paralili_stiriksi_table = self.construct_paralili_stiriksi_table()
        finished = datetime.now().replace(microsecond=0)
        logger.info("(aggregator) completed construct_paralili_stiriksi_table in just '%s'", finished - started)

        workbook = Workbook()

        # register normal sheet header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, size=9)
        bd = Side(style='thick', color="000000")
        header_style.border = Border(bottom=bd)
        header_style.alignment = Alignment(textRotation=90, wrapText=True, horizontal="center", vertical="center")
        workbook.add_named_style(header_style)

        # register "description" sheet header style
        header_style = NamedStyle(name="header_style_description")
        header_style.font = Font(bold=True, size=9)
        bd = Side(style='thick', color="000000")
        header_style.border = Border(bottom=bd)
        header_style.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
        workbook.add_named_style(header_style)

        # Γενικής Αγωγής

        ws = workbook.active
        ws.title = 'Γενικής Αγωγής'

        if len(self.gesTable) > 1:

            filter_headers(self.gesTable[0])

            for row in self.gesTable:
                ws.append(row)

            style_ws(ws)

        # Γενικής Αγωγής Παρατηρήσεις
        ws = workbook.create_sheet(title='Γενικής Αγωγής Παρατηρήσεις')

        report_rows = self.construct_ges_description_table()
        if len(report_rows) > 1:

            for row in report_rows:
                ws.append(row)

            style_description_ws(ws)

        # Τμήματα Ένταξης
        ws = workbook.create_sheet(title='Τμήματα Ένταξης')
        if len(tmimata_entakseis_table) > 1:

            filter_headers(tmimata_entakseis_table[0])

            for row in tmimata_entakseis_table:
                ws.append(row)

            style_ws(ws)

        # Τμήματα Ένταξης Παρατηρήσεις
        ws = workbook.create_sheet(title='Τμήματα Ένταξης Παρατηρήσεις')

        report_rows = self.construct_tmimata_entaksis_description_table()
        if len(report_rows) > 1:

            for row in report_rows:
                ws.append(row)

            style_description_ws(ws)

        # Παράλληλη Στήριξη
        ws = workbook.create_sheet(title='Παράλληλη Στήριξη')
        if len(paralili_stiriksi_table) > 1:

            filter_headers(paralili_stiriksi_table[0])

            for row in paralili_stiriksi_table:
                ws.append(row)

            style_ws(ws)

        # Παράλληλη Στήριξη Παρατηρήσεις
        ws = workbook.create_sheet(title='Παράλληλη Στήριξη Παρατηρήσεις')

        report_rows = self.construct_paralili_stiriksi_description_table()
        if len(report_rows) > 1:

            for row in report_rows:
                ws.append(row)

            style_description_ws(ws)

        # Ειδικής Αγωγής
        ws = workbook.create_sheet(title='Ειδικής Αγωγής')
        if len(self.sesTable) > 1:

            filter_headers(self.sesTable[0])

            for row in self.sesTable:
                ws.append(row)

            style_ws(ws)

        # Ειδικής Αγωγής Παρατηρήσεις
        ws = workbook.create_sheet(title='Ειδικής Αγωγής Παρατηρήσεις')

        report_rows = self.construct_ses_description_table()
        if len(report_rows) > 1:

            for row in report_rows:
                ws.append(row)

            style_description_ws(ws)

        # Υπόλοιπα
        ws = workbook.create_sheet(title='Υπόλοιπα')
        if len(self.msTable) > 1:

            filter_headers(self.msTable[0])

            for row in self.msTable:
                ws.append(row)

            style_ws(ws)

        # Υπόλοιπα Παρατηρήσεις
        ws = workbook.create_sheet(title='Υπόλοιπα Παρατηρήσεις')

        report_rows = self.construct_misc_description_table()
        if len(report_rows) > 1:

            for row in report_rows:
                ws.append(row)

            style_description_ws(ws)

        return workbook


@login_required
@user_passes_test(check_user_is_superuser)
@csrf_protect
def add_specialties(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["file"]

                workbook = load_workbook(excel_file)
                worksheet = workbook.active
                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        if cell.value is None:
                            text = ""
                        else:
                            text = str(cell.value).strip()

                        row_data.append(text)
                    excel_data.append(row_data)

                for row in excel_data[1:]:
                    
                    code = row[0]
                    lectic = row[1]

                    try:
                        label = row[2].strip() if row[2] is not None and len(row[2].strip()) > 0 else None
                    except:
                        label = None

                    try:
                        ordering = int(row[3])
                    except:
                        ordering = 0
                    
                    try:
                        is_active = str2bool(row[4])
                    except:
                        is_active = True

                    try:
                        # specialty exists
                        existing_specialty: Specialty = Specialty.objects.get(code=code)
                        existing_specialty.lectic = lectic
                        existing_specialty.label = label
                        existing_specialty.ordering = ordering
                        existing_specialty.active = is_active
                        existing_specialty.save()
                        logger.info("specialty '%s' has been successfully updated", code)
                    except Specialty.DoesNotExist:
                        # specialty does not exist, let's create one
                        Specialty.objects.create(code=code, lectic=lectic, label=label, ordering=ordering, active=is_active)
                        logger.info("specialty '%s' has been just created", code)
                        

                    
            except:
                messages.warning(request,
                                 "Κάτι πήγε λάθος... Μάλλον οι εγγραφές που μόλις φόρτωσες υπήρχαν ήδη στον πίνακα.")
                logger.exception("Something went wrong!")

        return render(request, 'excels/add_specialties.html', {'form': form, 'excel_data': excel_data})
    else:
        form = UploadFileForm()
        return render(request, 'excels/add_specialties.html', {'form': form})


def get_or_create_school_group(school_group_number: int) -> SchoolGroup:

    school_group_name = f"{school_group_number}η ομάδα"  # sorry about that
    try:
        return SchoolGroup.objects.get(name=school_group_name)
    except SchoolGroup.DoesNotExist:
        # school group does not exist, go ahead and create one
        return SchoolGroup.objects.create(name=school_group_name)


def get_or_create_sibling_school(school_ministry_code: str) -> School:
    try:
        return School.objects.get(ministry_code=school_ministry_code)
    except School.DoesNotExist:
        # school does not exist. Create a sibling school with radnom data that will
        # be fixed anyway, during the school upload process

        return School.objects.create(
            ministry_code=school_ministry_code,
            email=f'{get_random_string(length=12)}@example.gr',
            name=get_random_string(length=16)
        )


@login_required
@user_passes_test(check_user_is_superuser)
@csrf_protect
def add_schools(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["file"]
                purge_existing_schools = form.cleaned_data.get('purge_existing_schools', False)

                workbook = load_workbook(excel_file)
                worksheet = workbook.active
                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        if cell.value is None:
                            text = ""
                        else:
                            text = str(cell.value).strip()

                        row_data.append(text)
                    excel_data.append(row_data)

                with transaction.atomic():

                    if purge_existing_schools:
                        School.objects.all().delete()
                        Profile.objects.all().update(verified=False)
                        User.objects.all().update(last_name='')

                    for row in excel_data[1:]:
                        ministry_code = row[0]
                        sibling_school_ministry_code = row[1]
                        school_group_number = row[2]
                        neighboring_groups = row[3]
                        school_type = row[4]
                        school_variant = row[5]
                        myschool_school_name = row[6]
                        school_name = row[7]
                        email = row[8]
                        principal = row[9]
                        phone = row[10]
                        address = row[11]

                        if len(school_group_number) > 0:
                            school_group = get_or_create_school_group(int(school_group_number))
                        else:
                            school_group = None

                        if school_group is not None:
                            # handle neighboring groups of group
                            neighboring_groups_list = list()
                            group_tag = school_group.neighboring_tag

                            if len(neighboring_groups) > 0:
                                neighboring_groups_list = [get_or_create_school_group(x.strip()) for x in
                                                           neighboring_groups.split(',')]
                            else:
                                neighboring_groups_list = list()

                            if group_tag is None:
                                for neighboring_group in neighboring_groups_list:
                                    if neighboring_group.neighboring_tag is not None:
                                        group_tag = neighboring_group.neighboring_tag
                                        break

                            if group_tag is None:
                                group_tag = get_random_string(5)

                            school_group.neighboring_tag = group_tag
                            school_group.save()

                            for neighboring_group in neighboring_groups_list:
                                neighboring_group.neighboring_tag = group_tag
                                neighboring_group.save()

                        try:
                            existing_school: School = School.objects.get(email=email)
                        except School.DoesNotExist:
                            existing_school = None

                        # if we failed to fetch existing school by email, give it another try
                        # with ministry code. We need this due to the sibling school code as well
                        if existing_school is None:
                            try:
                                existing_school: School = School.objects.get(ministry_code=ministry_code)
                            except School.DoesNotExist:
                                existing_school = None

                        if len(sibling_school_ministry_code) > 0:
                            sibling_school = get_or_create_sibling_school(school_ministry_code=sibling_school_ministry_code)
                        else:
                            sibling_school = None

                        if existing_school is not None:
                            # we are updating
                            existing_school.email = email
                            existing_school.name = school_name
                            existing_school.school_group = school_group
                            existing_school.school_type = school_type
                            existing_school.school_variant = school_variant
                            existing_school.sibling_school = sibling_school
                            existing_school.myschool_name = myschool_school_name
                            existing_school.address = address
                            existing_school.principal = principal
                            existing_school.phone = phone
                            existing_school.ministry_code = ministry_code
                            existing_school.save()
                        else:
                            School.objects.create(
                                email=email,
                                name=school_name,
                                school_group=school_group,
                                school_type=school_type,
                                school_variant=school_variant,
                                sibling_school=sibling_school,
                                myschool_name=myschool_school_name,
                                address=address,
                                principal=principal,
                                phone=phone,
                                ministry_code=ministry_code
                            )

                    reconnect_nv_users_to_schools()
            except:
                messages.warning(request,
                                 "Κάτι πήγε λάθος... Μάλλον οι εγγραφές που μόλις φόρτωσες υπήρχαν ήδη στον πίνακα.")
                logger.exception("Something went wrong!")

        return render(request, 'excels/add_schools.html', {'form': form, 'excel_data': excel_data})
    else:
        form = UploadFileForm()
        return render(request, 'excels/add_schools.html', {'form': form})


@login_required
@user_passes_test(check_user_is_superuser)
def excel_entries(request):

    entries = Entry.objects.all().order_by('specialty')

    data = list()
    header = [
        'Σχολείο', 'Ειδικότητα', 'Είδος', 'Τύπος', 'Ώρες',
        'Παρατηρήσεις', 'Χρονική σήμανση'
    ]
    data.append(header)
    for entry in entries:
        row = [
            entry.owner.last_name, entry.specialty.code, entry.type,
            str(EntryVariantType(entry.variant).label), entry.hours,
            entry.description, entry.date_time,
        ]
        data.append(row)

    return ExcelResponse(data, 'entries')


@login_required
@user_passes_test(check_user_is_superuser)
def excel_history(request):

    entries = HistoryEntry.objects.all().order_by('owner', 'specialty')

    data = list()
    header = [
        'Σχολείο', 'Ειδικότητα', 'Είδος', 'Τύπος', 'Ώρες',
        'Παρατηρήσεις', 'Χρονική σήμανση'
    ]
    data.append(header)
    for entry in entries:
        row = [
            entry.owner.last_name, entry.specialty.code, entry.type,
            str(EntryVariantType(entry.variant).label), entry.hours,
            entry.description, entry.date_time
        ]
        data.append(row)

    return ExcelResponse(data, 'history')


@login_required
@user_passes_test(check_user_is_superuser)
def excel_user_history(request):
    entries = HistoryEntry.objects.filter(owner=request.user)

    data = list()
    header = ['Σχολείο', 'Ειδικότητα', 'Ώρες', 'Είδος', 'Παρατηρήσεις', 'Χρονική σήμανση']
    data.append(header)
    for entry in entries:
        row = [entry.owner.last_name, entry.specialty.code, entry.hours, entry.type,
               entry.description, entry.date_time]
        data.append(row)

    return ExcelResponse(data, 'user_history')


def reconnect_users_to_schools(user=None):
    # type: (User) -> List[Profile]
    if user is None:
        # we are working for all users
        profiles = Profile.objects.all()  # type: list[Profile]
    else:
        profiles = [user.profile, ]  # type: list[Profile]

    # store the associated (actually processed) users/profile in a list
    associated_users = list()

    for profile in profiles:
        try:
            school = School.objects.get(email=profile.user.email)  # type: School

            profile.user.last_name = school.name
            profile.user.save()

            profile.verified = True
            profile.school = school
            profile.save()

            associated_users.append(profile)
        except School.DoesNotExist:
            pass

    return associated_users


def reconnect_nv_users_to_schools():
    profiles = Profile.objects.all()  # type: list[Profile]

    for profile in profiles:
        try:
            school = School.objects.get(email=profile.user.email)  # type: School

            profile.user.last_name = school.name
            profile.user.save()
            profile.school = school
            profile.verified = True
            profile.save()
        except School.DoesNotExist:
            pass


@login_required
@user_passes_test(check_user_is_superuser)
def excel_aggregated_entries(request):

    report = AggregatedEntriesReport()
    workbook = report.get_workbook()

    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=kena_aggregated.xlsx'
    return response
