from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension
from openpyxl.utils import get_column_letter

from django.utils import timezone
from django.db import transaction
from django.db.models.query import QuerySet
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test, login_required
from django.http import HttpResponseRedirect, HttpRequest, HttpResponse
from django.urls import reverse, reverse_lazy
from django.contrib.auth.models import User
from django.views.generic import (
    ListView,
    CreateView,
    UpdateView,
    DeleteView
)
from .models import Covid19Entry, Covid19EntryEventHistory
from .forms import Covid19EntryCreateForm, Covid19EntryUpdateForm
from users.models import Profile
from vacancies.commons import EntryHistoryEventType
from vacancies.utils.view import UserIsAssociatedWithASchoolTestMixin
from vacancies.utils.permissions import check_user_is_superuser


class Covid19EntryListView(LoginRequiredMixin, UserIsAssociatedWithASchoolTestMixin, ListView):
    model = Covid19Entry
    context_object_name = 'entries'
    paginate_by = 10

    def get_queryset(self):

        user_profile: Profile = self.request.user.profile
        base_qs = super(Covid19EntryListView, self).get_queryset()
        return base_qs.filter(school=user_profile.school, deleted_on__isnull=True).order_by('specialty__ordering')


class Covid19EntryCreateView(LoginRequiredMixin, UserIsAssociatedWithASchoolTestMixin, CreateView):
    model = Covid19Entry
    form_class = Covid19EntryCreateForm

    success_url = reverse_lazy('covid19:covid19_entries')

    def get_form_kwargs(self):
        kwargs = super(Covid19EntryCreateView, self).get_form_kwargs()
        # https://stackoverflow.com/questions/32260785/django-validating-unique-together-constraints-in-a-modelform-with-excluded-fiel
        current_user = self.request.user
        user_profile = current_user.profile
        kwargs['instance'] = Covid19Entry(created_by=current_user, school=user_profile.school)
        kwargs['user'] = current_user

        return kwargs

    def form_valid(self, form):
        with transaction.atomic():
            
            created_datetime = timezone.now()

            current_user: User = self.request.user
            current_user.profile.status = False
            current_user.profile.status_time = None
            current_user.profile.save()
            
            covid19entry: Covid19Entry = form.instance
            covid19entry.save()
                        
            Covid19EntryEventHistory.objects.create(
                covid_entry=covid19entry,
                created_by=self.request.user,
                event_datetime=created_datetime,
                event_type=EntryHistoryEventType.HISTORY_EVENT_INSERT,
                event_description=f"H εγγραφή '{covid19entry}' καταχωρήθηκε"
            )

            messages.success(self.request, f'To κενό COVID-19 \'{covid19entry}\' καταχωρήθηκε με επιτυχία')

            return super().form_valid(form)


class Covid19EntryDeleteView(LoginRequiredMixin, UserIsAssociatedWithASchoolTestMixin, DeleteView):

    model = Covid19Entry

    def get_queryset(self):
        base_qs = super(Covid19EntryDeleteView, self).get_queryset()
        return base_qs.filter(school=self.request.user.profile.school, deleted_on__isnull=True)

    def delete(self, *args, **kwargs):
        with transaction.atomic():
            entry_to_delete: Covid19Entry = self.get_object()
            deleted_datetime = timezone.now()
            self.request.user.profile.status = False
            self.request.user.profile.status_time = None
            self.request.user.profile.save()

            Covid19EntryEventHistory.objects.create(
                covid_entry=entry_to_delete,
                created_by=self.request.user,
                event_datetime=deleted_datetime,
                event_type=EntryHistoryEventType.HISTORY_EVENT_DELETE,
                event_description=f"H εγγραφή '{entry_to_delete}' διαγράφηκε"
            )

            entry_to_delete.deleted_on = deleted_datetime
            entry_to_delete.save()

            messages.success(self.request, "Η διαγραφή του κενού COVID-19 ήταν επιτυχής")

        return HttpResponseRedirect(reverse('covid19:covid19_entries'))


class Covid19EntryUpdateView(LoginRequiredMixin, UserIsAssociatedWithASchoolTestMixin, UpdateView):
    model = Covid19Entry
    form_class = Covid19EntryUpdateForm
    context_object_name = 'entry'
    template_name_suffix = '_update_form'

    success_url = reverse_lazy('covid19:covid19_entries')

    def get_queryset(self):
        base_qs = super(Covid19EntryUpdateView, self).get_queryset()
        return base_qs.filter(school=self.request.user.profile.school, deleted_on__isnull=True)

    def form_valid(self, form):

        updated_datetime = timezone.now()

        original_data = self.get_object()
        updated_data: Covid19Entry = form.instance
        updated_data.updated_on = updated_datetime

        self.request.user.profile.status = False
        self.request.user.profile.status_time = None
        self.request.user.profile.save()

        Covid19EntryEventHistory.objects.create(
            covid_entry=original_data,
            created_by=self.request.user,
            event_datetime=updated_datetime,
            event_type=EntryHistoryEventType.HISTORY_EVENT_UPDATE,
            event_description=f"H εγγραφή '{original_data}' τροποποιήθηκε σε '{updated_data}'"
        )

        messages.success(self.request, "Η ενημέρωση του κενού COVID-19 ήταν επιτυχής")

        return super().form_valid(form)


@login_required
@user_passes_test(check_user_is_superuser)
def export_covid19_entries(request: HttpRequest):

    entries: QuerySet[Covid19Entry] = Covid19Entry.objects.filter(deleted_on__isnull=True).order_by('school',
                                                                                                    'specialty__ordering')

    now_datetime_string = timezone.now().strftime("%d-%m-%Y")

    workbook = Workbook()

    # register normal sheet header style
    header_style = NamedStyle(name="header_style")
    header_style.fill = PatternFill("solid", fgColor="DDDDDD")

    header_style.font = Font(bold=True, size=9)
    bd = Side(style='thick', color="000000")
    header_style.border = Border(bottom=bd)
    header_style.alignment = Alignment(textRotation=0, wrapText=True, horizontal="center", vertical="center")
    workbook.add_named_style(header_style)

    ws = workbook.active
    ws.title = 'Κενά COVID-19'
    print(now_datetime_string)

    header = [

        'Σχολείο', 'Επώνυμο Εκπαιδευτικού', 'Όνομα Εκπαιδευτικού', 'Ειδικότητα',
        'Ώρες', 'Ημ/νια Ασθένειας', 'Ημ/νια Επιστροφής', 'Ημ/νια Καταχώρισης',
        'Ημ/νια Αλλαγής', 'Σχόλια', '#'
    ]
    ws.append(header)

    entries_that_are_not_finalized = list()
    for entry in entries:

        # remember which entry belongs to a school that has not confirmed yet
        # data

        if entry.created_by.profile.status is False:
            entries_that_are_not_finalized.append(entry.pk)

        entry_created_on_str = entry.created_on.strftime('%d-%m-%Y %H:%M:%S') if entry.created_on else ''
        entry_updated_on_str = entry.updated_on.strftime('%d-%m-%Y %H:%M:%S') if entry.updated_on else ''

        row = [
            entry.school.name, entry.teacher_surname, entry.teacher_name,
            f'{entry.specialty.code} - {entry.specialty.lectic}', entry.hours,
            entry.illness_started, entry.illness_end_estimation, entry_created_on_str,
            entry_updated_on_str, entry.comments, entry.pk
        ]
        ws.append(row)

    # style the workbook

    # style header
    for col in range(ws.min_column, ws.max_column + 1):
        ws[get_column_letter(col) + '1'].style = 'header_style'

    # update column dimensions
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        if col == 1:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=23)
        elif col == 2:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
        elif col == 3:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)
        elif col == 4:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=34)
        elif col == 5:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=5)
        elif col == 6:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
        elif col == 7:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
        elif col == 8:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        elif col == 9:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        elif col == 10:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=40)
        else:
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=14)

    ws.column_dimensions = dim_holder

    # update 1st row height
    ws.row_dimensions[1] = RowDimension(ws, height=30)

    fill = PatternFill("solid", fgColor="FF0000")

    for row in range(ws.min_row, ws.max_row + 1):
        test_column = ws[get_column_letter(ws.max_column) + str(row)]
        if test_column.value in entries_that_are_not_finalized:
            for col in range(ws.min_column, ws.max_column + 1):
                ws[get_column_letter(col) + str(row)].fill = fill

    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=covid19-{now_datetime_string}.xlsx'
    return response



